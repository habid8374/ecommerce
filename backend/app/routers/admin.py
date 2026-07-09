"""Admin panel: order management, dashboard metrics, and customers."""
import io
from datetime import datetime, timedelta, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from ..database import get_db
from ..deps import get_current_admin
from ..models import (
    CustomerSummary,
    DashboardStats,
    Order,
    OrderStatus,
    OrderStatusUpdate,
    PaymentStatus,
    UserPublic,
    _now,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])

PROJECT = {"_id": 0}


# --- Orders ---------------------------------------------------------------
@router.get("/orders")
async def list_orders(
    status_filter: Optional[OrderStatus] = Query(None, alias="status"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    _: UserPublic = Depends(get_current_admin),
):
    db = get_db()
    query: dict = {}
    if status_filter:
        query["status"] = status_filter.value
    total = await db.orders.count_documents(query)
    cursor = (
        db.orders.find(query, PROJECT)
        .sort("created_at", -1)
        .skip((page - 1) * page_size)
        .limit(page_size)
    )
    items = [Order(**o) for o in await cursor.to_list(page_size)]
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, _: UserPublic = Depends(get_current_admin)):
    db = get_db()
    order = await db.orders.find_one({"id": order_id}, PROJECT)
    if not order:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Pedido no encontrado")
    return Order(**order)


@router.patch("/orders/{order_id}/status", response_model=Order)
async def update_order_status(
    order_id: str, body: OrderStatusUpdate, _: UserPublic = Depends(get_current_admin)
):
    db = get_db()
    result = await db.orders.update_one(
        {"id": order_id},
        {"$set": {"status": body.status.value, "updated_at": _now()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Pedido no encontrado")
    order = await db.orders.find_one({"id": order_id}, PROJECT)
    # Notify the customer of the new status (best-effort).
    try:
        from ..services.email import send_status_changed

        await send_status_changed(order, body.status.value)
    except Exception:  # noqa: BLE001
        pass
    return Order(**order)


@router.patch("/orders/{order_id}/confirm-payment", response_model=Order)
async def confirm_payment(order_id: str, _: UserPublic = Depends(get_current_admin)):
    """Manually mark a payment as verified (e.g. bank transfer / cash), in
    addition to the automatic Wompi webhook. Decrements stock once, idempotent."""
    from .orders import mark_order_paid

    db = get_db()
    order = await db.orders.find_one({"id": order_id}, PROJECT)
    if not order:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Pedido no encontrado")
    await mark_order_paid(db, order, order.get("wompi_transaction_id") or "MANUAL")
    order = await db.orders.find_one({"id": order_id}, PROJECT)
    return Order(**order)


@router.delete("/orders/{order_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_order(order_id: str, _: UserPublic = Depends(get_current_admin)):
    """Delete any order. If it had been paid (stock already decremented), the
    stock is restored so inventory stays correct."""
    db = get_db()
    order = await db.orders.find_one({"id": order_id}, PROJECT)
    if not order:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Pedido no encontrado")
    if order.get("payment_status") == PaymentStatus.approved.value:
        for item in order.get("items", []):
            await db.products.update_one(
                {"id": item["product_id"]},
                {"$inc": {"stock": int(item["quantity"])}},
            )
    await db.orders.delete_one({"id": order_id})


@router.get("/orders/pending-count")
async def pending_payment_count(_: UserPublic = Depends(get_current_admin)):
    """Orders awaiting payment verification — used for the live admin badge."""
    db = get_db()
    n = await db.orders.count_documents({"payment_status": PaymentStatus.pending.value})
    return {"pending_payment": n}


# --- Dashboard ------------------------------------------------------------
@router.get("/stats", response_model=DashboardStats)
async def dashboard_stats(_: UserPublic = Depends(get_current_admin)):
    db = get_db()

    revenue_agg = await db.orders.aggregate(
        [
            {"$match": {"payment_status": PaymentStatus.approved.value}},
            {"$group": {"_id": None, "total": {"$sum": "$total"}}},
        ]
    ).to_list(1)
    revenue = revenue_agg[0]["total"] if revenue_agg else 0

    recent_cursor = db.orders.find({}, PROJECT).sort("created_at", -1).limit(5)
    recent = [Order(**o) for o in await recent_cursor.to_list(5)]

    return DashboardStats(
        revenue=revenue,
        orders_total=await db.orders.count_documents({}),
        orders_pending=await db.orders.count_documents({"status": OrderStatus.pending.value}),
        orders_paid=await db.orders.count_documents(
            {"payment_status": PaymentStatus.approved.value}
        ),
        customers_total=await db.users.count_documents({"role": "customer"}),
        products_total=await db.products.count_documents({}),
        low_stock=await db.products.count_documents({"stock": {"$lte": 5}}),
        recent_orders=recent,
    )


# --- Analytics (professional dashboard) -----------------------------------
@router.get("/analytics")
async def analytics(days: int = 30, _: UserPublic = Depends(get_current_admin)):
    db = get_db()
    since = datetime.now(timezone.utc) - timedelta(days=days)
    paid = {"payment_status": PaymentStatus.approved.value}

    # Revenue & orders per day.
    sales = await db.orders.aggregate([
        {"$match": {**paid, "created_at": {"$gte": since}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$created_at"}},
            "revenue": {"$sum": "$total"},
            "orders": {"$sum": 1},
        }},
        {"$sort": {"_id": 1}},
    ]).to_list(1000)
    sales_series = [{"date": s["_id"], "revenue": s["revenue"], "orders": s["orders"]} for s in sales]

    # Best sellers by units.
    top = await db.orders.aggregate([
        {"$match": paid},
        {"$unwind": "$items"},
        {"$group": {
            "_id": "$items.name",
            "qty": {"$sum": "$items.quantity"},
            "revenue": {"$sum": "$items.subtotal"},
        }},
        {"$sort": {"qty": -1}},
        {"$limit": 8},
    ]).to_list(8)
    top_products = [{"name": t["_id"], "qty": t["qty"], "revenue": t["revenue"]} for t in top]

    # Units sold per product id -> low rotation.
    sold = await db.orders.aggregate([
        {"$match": paid},
        {"$unwind": "$items"},
        {"$group": {"_id": "$items.product_id", "qty": {"$sum": "$items.quantity"}}},
    ]).to_list(10000)
    sold_map = {s["_id"]: s["qty"] for s in sold}
    products = await db.products.find({"active": True}, PROJECT).to_list(10000)
    low = sorted(products, key=lambda p: sold_map.get(p["id"], 0))[:8]
    low_rotation = [
        {"name": p["name"], "qty": sold_map.get(p["id"], 0), "stock": p.get("stock", 0)}
        for p in low
    ]

    # Revenue by category.
    by_cat = await db.orders.aggregate([
        {"$match": paid},
        {"$unwind": "$items"},
        {"$lookup": {
            "from": "products", "localField": "items.product_id",
            "foreignField": "id", "as": "p",
        }},
        {"$group": {
            "_id": {"$ifNull": [{"$arrayElemAt": ["$p.category", 0]}, "otros"]},
            "revenue": {"$sum": "$items.subtotal"},
        }},
        {"$sort": {"revenue": -1}},
    ]).to_list(50)
    revenue_by_category = [{"category": c["_id"], "revenue": c["revenue"]} for c in by_cat]

    status_breakdown = {
        st.value: await db.orders.count_documents({"status": st.value}) for st in OrderStatus
    }

    revenue_total = sum(s["revenue"] for s in sales_series)
    orders_total = sum(s["orders"] for s in sales_series)
    return {
        "days": days,
        "sales_series": sales_series,
        "top_products": top_products,
        "low_rotation": low_rotation,
        "revenue_by_category": revenue_by_category,
        "status_breakdown": status_breakdown,
        "period_revenue": revenue_total,
        "period_orders": orders_total,
        "avg_ticket": round(revenue_total / orders_total) if orders_total else 0,
    }


# --- Customers ------------------------------------------------------------
async def _customer_summaries(db) -> list[CustomerSummary]:
    users = await db.users.find({"role": "customer"}, {"_id": 0, "password": 0}).sort(
        "created_at", -1
    ).to_list(5000)

    summaries: list[CustomerSummary] = []
    for u in users:
        agg = await db.orders.aggregate(
            [
                {
                    "$match": {
                        "user_id": u["id"],
                        "payment_status": PaymentStatus.approved.value,
                    }
                },
                {"$group": {"_id": None, "spent": {"$sum": "$total"}, "count": {"$sum": 1}}},
            ]
        ).to_list(1)
        spent = agg[0]["spent"] if agg else 0
        count = agg[0]["count"] if agg else 0

        # Phone/city: prefer the profile phone, fall back to the latest order.
        latest = await db.orders.find({"user_id": u["id"]}, PROJECT).sort(
            "created_at", -1
        ).limit(1).to_list(1)
        addr = (latest[0].get("shipping_address") if latest else None) or {}
        phone = u.get("phone") or addr.get("phone", "")
        city = addr.get("city", "")

        summaries.append(
            CustomerSummary(
                id=u["id"],
                name=u["name"],
                email=u["email"],
                phone=phone,
                city=city,
                created_at=u["created_at"],
                orders_count=count,
                total_spent=spent,
            )
        )
    return summaries


@router.get("/customers", response_model=list[CustomerSummary])
async def list_customers(_: UserPublic = Depends(get_current_admin)):
    return await _customer_summaries(get_db())


def _fmt_date(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value or "")


@router.get("/customers/export")
async def export_customers(_: UserPublic = Depends(get_current_admin)):
    """Download an .xlsx with all customers and their purchases (for marketing)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db()
    summaries = await _customer_summaries(db)

    wb = Workbook()

    # Sheet 1: customers
    ws = wb.active
    ws.title = "Clientes"
    headers = ["Nombre", "Email", "Teléfono", "Ciudad", "Registrado", "N° Pedidos", "Total gastado (COP)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for c in summaries:
        ws.append([
            c.name, c.email, c.phone, c.city, _fmt_date(c.created_at),
            c.orders_count, c.total_spent,
        ])

    # Sheet 2: every order (their purchases)
    ws2 = wb.create_sheet("Pedidos")
    ws2.append([
        "Pedido", "Cliente", "Email", "Teléfono", "Ciudad", "Fecha",
        "Estado", "Pago", "Artículos", "Total (COP)",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    orders = await db.orders.find({}, PROJECT).sort("created_at", -1).to_list(50000)
    for o in orders:
        addr = o.get("shipping_address") or {}
        items = "; ".join(f"{it['quantity']}x {it['name']}" for it in o.get("items", []))
        ws2.append([
            o["id"][:8], addr.get("full_name", ""), o.get("customer_email", ""),
            addr.get("phone", ""), addr.get("city", ""), _fmt_date(o.get("created_at")),
            o.get("status", ""), o.get("payment_status", ""), items, o.get("total", 0),
        ])

    # Auto-ish column widths
    for sheet in (ws, ws2):
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"clientes_grafibless_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
