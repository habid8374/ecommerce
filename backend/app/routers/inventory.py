"""Inventory module: stock valuation, movements (kardex), adjustments,
per-category IVA, and an Excel report."""
import io
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from ..database import get_db
from ..deps import get_current_admin
from ..models import MovementCreate, TaxByCategory, UserPublic, _now
from ..services.inventory import apply_movement

router = APIRouter(prefix="/api/admin/inventory", tags=["admin"])

PROJECT = {"_id": 0}


def _row(p: dict) -> dict:
    stock = int(p.get("stock", 0) or 0)
    cost = int(p.get("cost", 0) or 0)
    price = int(p.get("price", 0) or 0)
    threshold = int(p.get("low_stock_threshold", 5) or 0)
    is_service = bool(p.get("is_service"))
    return {
        "id": p["id"],
        "name": p.get("name", ""),
        "sku": p.get("sku", ""),
        "barcode": p.get("barcode", ""),
        "category": p.get("category", ""),
        "stock": stock,
        "cost": cost,
        "price": price,
        "tax_rate": int(p.get("tax_rate", 0) or 0),
        "low_stock_threshold": threshold,
        "is_service": is_service,
        "active": bool(p.get("active", True)),
        "cost_value": stock * cost,
        "retail_value": stock * price,
        "low": (not is_service) and stock <= threshold,
        "out": (not is_service) and stock <= 0,
    }


@router.get("")
async def inventory_overview(_: UserPublic = Depends(get_current_admin)):
    db = get_db()
    products = await db.products.find({}, PROJECT).sort("name", 1).to_list(10000)
    rows = [_row(p) for p in products]
    goods = [r for r in rows if not r["is_service"]]
    summary = {
        "skus": len(goods),
        "units": sum(r["stock"] for r in goods),
        "cost_value": sum(r["cost_value"] for r in goods),
        "retail_value": sum(r["retail_value"] for r in goods),
        "low_count": sum(1 for r in goods if r["low"]),
        "out_count": sum(1 for r in goods if r["out"]),
        "services": sum(1 for r in rows if r["is_service"]),
    }
    return {"summary": summary, "items": rows}


@router.get("/movements")
async def list_movements(
    product_id: str = Query(default=""),
    limit: int = Query(default=200, ge=1, le=2000),
    _: UserPublic = Depends(get_current_admin),
):
    db = get_db()
    q = {"product_id": product_id} if product_id else {}
    docs = await db.inventory_movements.find(q, PROJECT).sort("created_at", -1).to_list(limit)
    return docs


@router.post("/apply-tax")
async def apply_tax_to_category(
    body: TaxByCategory, admin: UserPublic = Depends(get_current_admin)
):
    """Bulk-set the IVA % for every product in a category."""
    db = get_db()
    cat = (body.category or "").strip().lower()
    res = await db.products.update_many(
        {"category": cat}, {"$set": {"tax_rate": body.tax_rate, "updated_at": _now()}}
    )
    return {"updated": res.modified_count, "category": cat, "tax_rate": body.tax_rate}


@router.get("/export")
async def export_inventory(_: UserPublic = Depends(get_current_admin)):
    """Download an .xlsx with the stock valuation and the movements (kardex)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    db = get_db()
    products = await db.products.find({}, PROJECT).sort("name", 1).to_list(10000)
    rows = [_row(p) for p in products]

    wb = Workbook()
    ws = wb.active
    ws.title = "Existencias"
    ws.append([
        "Producto", "SKU", "Código barras", "Categoría", "Tipo", "Stock", "Costo unit.",
        "Valor costo", "Precio", "Valor venta", "IVA %", "Mínimo", "Estado",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in rows:
        svc = r["is_service"]
        # Services don't hold stock: leave stock/valuation blank so the sheet
        # isn't padded with meaningless numbers, and don't repeat "Servicio"
        # in Estado (the Tipo column already says it).
        estado = "" if svc else ("Agotado" if r["out"] else ("Bajo" if r["low"] else "OK"))
        ws.append([
            r["name"], r["sku"], r["barcode"], r["category"], "Servicio" if svc else "Producto",
            "" if svc else r["stock"], r["cost"],
            "" if svc else r["cost_value"], r["price"],
            "" if svc else r["retail_value"], r["tax_rate"],
            "" if svc else r["low_stock_threshold"], estado,
        ])

    ws2 = wb.create_sheet("Movimientos")
    ws2.append([
        "Fecha", "Producto", "Tipo", "Cambio", "Stock anterior", "Stock nuevo",
        "Costo unit.", "Motivo", "Pedido",
    ])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    TYPE_ES = {"purchase": "Entrada", "sale": "Venta", "adjustment": "Ajuste", "return": "Devolución"}
    movements = await db.inventory_movements.find({}, PROJECT).sort("created_at", -1).to_list(50000)
    for m in movements:
        created = m.get("created_at")
        fecha = created.strftime("%Y-%m-%d %H:%M") if isinstance(created, datetime) else str(created or "")
        ws2.append([
            fecha, m.get("product_name", ""), TYPE_ES.get(m.get("type"), m.get("type", "")),
            m.get("change", 0), m.get("previous_stock", 0), m.get("new_stock", 0),
            m.get("unit_cost", 0), m.get("reason", ""), (m.get("order_id") or "")[:8],
        ])

    for sheet in (ws, ws2):
        for col in sheet.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"inventario_grafibless_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/{product_id}/movement")
async def create_movement(
    product_id: str, body: MovementCreate, admin: UserPublic = Depends(get_current_admin)
):
    """Register a manual entrada/salida/ajuste for a product."""
    db = get_db()
    product = await db.products.find_one({"id": product_id}, PROJECT)
    if not product:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Producto no encontrado")

    if body.kind == "in":
        change, mtype = body.quantity, "purchase"
    elif body.kind == "out":
        change, mtype = -body.quantity, "adjustment"
    elif body.kind == "set":
        change, mtype = body.quantity - int(product.get("stock", 0) or 0), "adjustment"
    else:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Tipo de movimiento inválido")

    mv = await apply_movement(
        db, product, change, mtype,
        reason=body.reason, unit_cost=body.unit_cost, created_by=admin.email,
    )
    return mv.model_dump()
